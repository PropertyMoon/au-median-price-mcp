"""
Victoria median house/unit prices — downloads the "Victorian Property Sales Report -
Median House by Suburb Quarterly" Excel file from data.vic.gov.au (CKAN API discovery)
and parses it into a suburb lookup.

The XLS has one sheet per dwelling type (House, Unit) with columns:
  Suburb | Municipality | Median | No. Transfers | Quarter | ...

Data is cached in-process for 24 hours; first request triggers download (~1-3 MB).
"""

import asyncio
import io
import time

import httpx
import openpyxl

_CKAN_API  = "https://discover.data.vic.gov.au/api/3/action/package_show"
_PACKAGE_IDS = [
    "victorian-property-sales-report-median-house-by-suburb-quarterly",
    # fallback slug variant
    "property-sales-statistics-quarterly-median-sale-price-suburb",
]
_FALLBACK_URL = (
    "https://discover.data.vic.gov.au/dataset/"
    "victorian-property-sales-report-median-house-by-suburb-quarterly"
)

_cache: dict = {}
_cache_ts: float = 0.0
_cache_lock = asyncio.Lock()


async def _find_xlsx_url() -> str:
    async with httpx.AsyncClient(timeout=15) as client:
        for pkg_id in _PACKAGE_IDS:
            try:
                r = await client.get(_CKAN_API, params={"id": pkg_id})
                if r.status_code != 200:
                    continue
                resources = r.json().get("result", {}).get("resources", [])
                xlsx = [
                    res for res in resources
                    if (res.get("format") or "").upper() in ("XLSX", "XLS")
                ]
                if xlsx:
                    xlsx.sort(key=lambda x: x.get("created", ""), reverse=True)
                    return xlsx[0]["url"]
            except Exception:
                continue
    return _FALLBACK_URL


def _parse_quarter(val) -> str | None:
    """Normalise quarter cell value to 'YYYY QN' string."""
    if not val:
        return None
    s = str(val).strip()
    # Already formatted: "2024 Q3", "Q3 2024"
    import re
    m = re.search(r"(\d{4})[\s\-/]?Q(\d)", s, re.I)
    if m:
        return f"{m.group(1)} Q{m.group(2)}"
    m = re.search(r"Q(\d)[\s\-/]?(\d{4})", s, re.I)
    if m:
        return f"{m.group(2)} Q{m.group(1)}"
    return s


def _col_indices(headers: list) -> dict:
    """Return a mapping of role → column index from a header row."""
    result = {}
    for i, h in enumerate(headers):
        if h is None:
            continue
        h_low = str(h).strip().lower()
        if "suburb" in h_low:
            result.setdefault("suburb", i)
        elif "median" in h_low and "price" not in h_low:
            result.setdefault("median", i)
        elif "median" in h_low and "price" in h_low:
            result.setdefault("median", i)
        elif "quarter" in h_low or "period" in h_low or "qtr" in h_low:
            result.setdefault("quarter", i)
        elif "transfer" in h_low or "sales" in h_low or "number" in h_low:
            result.setdefault("transfers", i)
    return result


async def _load() -> dict:
    url = await _find_xlsx_url()
    async with httpx.AsyncClient(timeout=90, follow_redirects=True) as client:
        r = await client.get(url)
        r.raise_for_status()

    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)

    # suburb → {median_house_price, median_unit_price, price_history_quarterly, data_period}
    house_data: dict[str, dict] = {}
    unit_data:  dict[str, dict] = {}

    for sheet_name in wb.sheetnames:
        sn_low = sheet_name.strip().lower()
        if "house" in sn_low:
            target = house_data
        elif "unit" in sn_low or "flat" in sn_low or "apt" in sn_low:
            target = unit_data
        else:
            continue

        ws = wb[sheet_name]
        cols: dict = {}
        latest_quarter = ""

        # suburb → list of (quarter_str, median_price)
        suburb_history: dict[str, list] = {}

        for row in ws.iter_rows(values_only=True):
            if not any(row):
                continue
            # Detect header row
            if not cols:
                if any(
                    cell is not None and "suburb" in str(cell).lower()
                    for cell in row
                ):
                    cols = _col_indices(list(row))
                continue

            if "suburb" not in cols or "median" not in cols:
                continue

            suburb_raw = row[cols["suburb"]]
            if not suburb_raw:
                continue
            suburb_key = str(suburb_raw).strip().upper()

            median_raw = row[cols["median"]]
            try:
                median = int(float(str(median_raw).replace(",", "").replace("$", "")))
                if median <= 0:
                    continue
            except (ValueError, TypeError):
                continue

            quarter_str = None
            if "quarter" in cols:
                quarter_str = _parse_quarter(row[cols["quarter"]])
            if not quarter_str:
                quarter_str = "latest"

            if quarter_str > latest_quarter:
                latest_quarter = quarter_str

            suburb_history.setdefault(suburb_key, []).append((quarter_str, median))

        # Build per-suburb summary: latest median + last 8 quarters history
        for suburb_key, history in suburb_history.items():
            history.sort(key=lambda x: x[0], reverse=True)
            latest_median = history[0][1] if history else None
            hist_list = [
                {"period": q, "median_price": p}
                for q, p in history[:8]
            ]
            hist_list.reverse()  # oldest first
            target[suburb_key] = {
                "median_price":    latest_median,
                "data_period":     history[0][0] if history else latest_quarter,
                "price_history":   hist_list,
            }

    wb.close()
    return {"house": house_data, "unit": unit_data}


async def _data() -> dict:
    global _cache, _cache_ts
    async with _cache_lock:
        if _cache and (time.time() - _cache_ts) < 86_400:
            return _cache
        _cache = await _load()
        _cache_ts = time.time()
        return _cache


async def get_vic_median(suburb: str) -> dict:
    suburb_upper = suburb.strip().upper()
    data = await _data()

    house = data["house"].get(suburb_upper)
    unit  = data["unit"].get(suburb_upper)

    if not house and not unit:
        return {
            "suburb":   suburb,
            "state":    "VIC",
            "coverage": "no_data",
            "message":  f"No median price records found for '{suburb}' in VIC dataset.",
        }

    result: dict = {
        "suburb":   suburb,
        "state":    "VIC",
        "coverage": "available",
        "data_source": "Victorian Property Sales Report — Median House by Suburb Quarterly (data.vic.gov.au)",
    }

    if house:
        result["median_house_price"]          = house["median_price"]
        result["data_period"]                  = house["data_period"]
        result["price_history_quarterly"]      = house["price_history"]

    if unit:
        result["median_unit_price"] = unit["median_price"]
        if "data_period" not in result:
            result["data_period"] = unit["data_period"]

    return result
